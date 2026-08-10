import json
import os
import re
import time
import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options


def get_excel_path():
    local_primary = r"C:\Users\sai\OneDrive\Desktop\flipkart\FK Walkthrough 10th aug.xlsx"
    github_primary = "FK Walkthrough 10th aug.xlsx"

    if os.path.exists(local_primary):
        return local_primary
    elif os.path.exists(github_primary):
        return github_primary
    elif os.path.exists("FK Walkthrough 10th aug"):
        return "FK Walkthrough 10th aug"
    else:
        raise FileNotFoundError("Could not locate 'FK Walkthrough 10th aug.xlsx'.")


def get_amazon_data(html_source):
    soup = BeautifulSoup(html_source, "html.parser")

    # 1. Check for CAPTCHA or Bot Detection Pages
    if "api-services-support@amazon.com" in html_source.lower() or "enter the characters you see below" in html_source.lower():
        print(" [Warning] Amazon CAPTCHA / Bot detection triggered!")

    # 2. Check if the product is Out of Stock / Currently Unavailable
    out_of_stock_elem = soup.select_one("#availability, .a-color-price, .a-size-medium.a-color-state")
    if out_of_stock_elem and "currently unavailable" in out_of_stock_elem.get_text().lower():
        return "N/A", "N/A", "N/A"

    if "currently unavailable." in soup.get_text().lower():
        buybox_price = soup.select_one("#buybox .a-price .a-offscreen, #corePrice_feature_div .a-offscreen")
        if not buybox_price:
            return "N/A", "N/A", "N/A"

    # 3. Extract Price strictly from the main buybox/core price area
    price = "N/A"
    price_selectors = [
        "#corePrice_feature_div .apex-pricetopay-value .a-offscreen",
        "#corePrice_feature_div .a-price .a-offscreen",
        "#corePriceDisplay_desktop_feature_div .a-price-whole",
        "#priceblock_ourprice",
        "#priceblock_dealprice",
        "#buybox .a-price .a-offscreen"
    ]

    for selector in price_selectors:
        price_elem = soup.select_one(selector)
        if price_elem and price_elem.text:
            cleaned = re.sub(r"[^\d.]", "", price_elem.text.replace(",", "").strip())
            if cleaned:
                price = cleaned
                break

    if price == "N/A":
        json_ld_scripts = soup.find_all("script", type="application/ld+json")
        for script in json_ld_scripts:
            if script.string:
                try:
                    data = json.loads(script.string)
                    if isinstance(data, list):
                        data = data[0]
                    if "offers" in data:
                        offers = data["offers"]
                        if isinstance(offers, list):
                            offers = offers[0]
                        p = offers.get("price") or offers.get("lowPrice")
                        if p and str(p) != "0":
                            price = str(p).replace(",", "").strip()
                            break
                except Exception:
                    continue

    # 4. Extract Rating strictly from main review section
    rating = "N/A"
    main_rating_container = soup.select_one(
        "#averageCustomerReviews, #acrPopover, #centerCol #averageCustomerReviews"
    )

    if main_rating_container:
        rating_text = main_rating_container.get_text()
        match = re.search(r"([0-5]\.[0-9]|[0-5])", rating_text)
        if match:
            rating = match.group(1)

    # 5. Extract Deal Tag
    tag = "N/A"
    if price != "N/A":
        tag_selectors = [
            "#dealBadgeSupportingText",
            "#dealBadge",
            ".dealBadge",
            "span.badge-text",
            "#dealBadge_feature_div span",
            "#editorialBadgeContainer"
        ]

        for selector in tag_selectors:
            tag_elem = soup.select_one(selector)
            if tag_elem:
                text = tag_elem.get_text(strip=True)
                if text and len(text) > 1:
                    tag = text
                    break

    return price, rating, tag


def process_amazon_home():
    excel_file = get_excel_path()
    wb = openpyxl.load_workbook(excel_file)

    # Robust Chrome Options to bypass Cloud Bot Detection
    chrome_options = Options()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--incognito")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1920,1080")
    
    # Custom Headers & Anti-Bot Spoofing Flags
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36")
    chrome_options.add_argument("--lang=en-US,en;q=0.9")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)

    print("Launching Anti-Detection Headless Chrome Driver...")
    driver = webdriver.Chrome(options=chrome_options)
    
    # Hide WebDriver presence from navigator JS object
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    target_sheet_name = None
    for name in wb.sheetnames:
        clean_name = name.strip().lower()
        if clean_name in ["amzon home", "amazon home", "amazonhome", "amzonhome"]:
            target_sheet_name = name
            break

    if not target_sheet_name:
        print("Error: Could not find 'amzon home' or 'amazon home' sheet in workbook.")
        driver.quit()
        return

    ws = wb[target_sheet_name]
    print(f"\n{'=' * 80}\nProcessing Sheet: '{target_sheet_name}' (Bypass Anti-Bot Mode)\n{'=' * 80}")

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]

    def get_col_idx(col_name, default_idx=-1):
        return headers.index(col_name) + 1 if col_name in headers else default_idx

    url_col_idx = get_col_idx("URL", 3)

    current_col_idx = get_col_idx("Current", -1)
    if current_col_idx == -1:
        current_col_idx = get_col_idx("Current Price", -1)

    rating_col_idx = get_col_idx("Ratings", -1)
    if rating_col_idx == -1:
        rating_col_idx = get_col_idx("Rating", -1)
    if rating_col_idx == -1:
        rating_col_idx = get_col_idx("Stars", -1)

    tag_col_idx = get_col_idx("Tag", -1)
    if tag_col_idx == -1:
        tag_col_idx = get_col_idx("Tag Name", -1)
    if tag_col_idx == -1:
        tag_col_idx = get_col_idx("Deal Tag", -1)

    if current_col_idx == -1:
        print(f"Error: Could not locate 'Current' column in headers: {headers}")
        driver.quit()
        return

    total_rows = ws.max_row

    for row in range(2, total_rows + 1):
        url_cell = ws.cell(row=row, column=url_col_idx).value
        url = str(url_cell).strip() if url_cell else ""

        if not url.startswith("http"):
            print(f"[{row-1}/{total_rows-1}] Row {row}: Invalid/Missing URL. Skipping.")
            continue

        asin_val = ws.cell(row=row, column=1).value or "N/A"

        try:
            driver.get(url)
            
            # Wait 7 seconds for dynamic rendering
            time.sleep(7.0)

            price, rating, tag = get_amazon_data(driver.page_source)

            # Update Price
            if price != "N/A":
                try:
                    price_val = float(price) if "." in price else int(price)
                except ValueError:
                    price_val = price
                ws.cell(row=row, column=current_col_idx, value=price_val)
                price_str = f"₹{price}"
            else:
                ws.cell(row=row, column=current_col_idx, value="N/A")
                price_str = "N/A"

            # Update Rating
            if rating_col_idx != -1:
                if rating != "N/A":
                    try:
                        rating_val = float(rating)
                    except ValueError:
                        rating_val = rating
                    ws.cell(row=row, column=rating_col_idx, value=rating_val)
                else:
                    ws.cell(row=row, column=rating_col_idx, value="N/A")

            # Update Tag
            if tag_col_idx != -1:
                ws.cell(row=row, column=tag_col_idx, value=tag)

            # Log Progress
            print(f"[{row-1}/{total_rows-1}] ASIN: {asin_val} | Price: {price_str} | Rating: {rating} | Tag: {tag}")

        except Exception as err:
            print(f"[{row-1}/{total_rows-1}] Row {row} Error processing URL: {err}")

    driver.quit()

    wb.save(excel_file)
    print("\n" + "=" * 80)
    print(f"Successfully updated all prices, ratings, and tags in '{target_sheet_name}'!")
    print(f"File saved at: {excel_file}")


if __name__ == "__main__":
    process_amazon_home()
